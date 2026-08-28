"""
SCRIPT 1: Basic Social Media Analytics
Demonstrates: Reading CSV, calculating metrics, exporting to Excel
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read the CSV file
df = pd.read_csv('social_media_data.csv')

# Calculate platform-wise statistics
platform_stats = df.groupby('platform').agg({
    'likes': 'sum',
    'comments': 'sum',
    'shares': 'sum',
    'followers_gained': 'sum',
    'engagement_rate': 'mean'
}).round(2)

platform_stats.columns = ['Total Likes', 'Total Comments', 'Total Shares', 'Followers Gained', 'Avg Engagement %']

# Identify top performing posts
top_posts = df.nlargest(5, 'engagement_rate')[['post_id', 'platform', 'content_type', 'likes', 'engagement_rate']]

# Create Excel workbook with styling
wb = Workbook()
ws = wb.active
ws.title = "Analytics"

# Title
ws['A1'] = "SOCIAL MEDIA ANALYTICS REPORT"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:E1')

# Platform Stats
ws['A3'] = "Platform Performance"
ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

row = 4
for col_num, header in enumerate(platform_stats.columns, 1):
    ws.cell(row=row, column=col_num).value = header
    ws.cell(row=row, column=col_num).font = Font(bold=True)

for row_num, (platform, row_data) in enumerate(platform_stats.iterrows(), 5):
    ws.cell(row=row_num, column=1).value = platform
    for col_num, value in enumerate(row_data, 2):
        ws.cell(row=row_num, column=col_num).value = value

# Top Posts Section
ws['A15'] = "Top 5 Performing Posts"
ws['A15'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A15'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

row = 16
for col_num, header in enumerate(top_posts.columns, 1):
    ws.cell(row=row, column=col_num).value = header
    ws.cell(row=row, column=col_num).font = Font(bold=True)

for row_num, (idx, row_data) in enumerate(top_posts.iterrows(), 17):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num).value = value

# Adjust column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18

wb.save('social_media_report.xlsx')
print("✅ Report generated: social_media_report.xlsx")
