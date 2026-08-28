"""
SCRIPT 2: Content Type Performance Analyzer
Demonstrates: Data filtering, conditional logic, creating recommendation columns
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Read data
df = pd.read_csv('social_media_data.csv')

# Analyze content type performance per platform
content_performance = df.groupby(['platform', 'content_type']).agg({
    'engagement_rate': ['mean', 'max'],
    'likes': 'mean',
    'followers_gained': 'sum'
}).round(2)

# Flatten multi-level columns for easier access
content_performance.columns = ['_'.join(col).strip() for col in content_performance.columns.values]

# Create recommendation column
def get_recommendation(platform, engagement):
    if engagement > 10:
        return "🔥 VIRAL - Keep producing!"
    elif engagement > 5:
        return "✅ Good - Continue strategy"
    else:
        return "⚠️ Needs improvement"

df['recommendation'] = df.apply(
    lambda row: get_recommendation(row['platform'], row['engagement_rate']),
    axis=1
)

# Best time to post (by hour)
best_hours = df.groupby('post_hour')['engagement_rate'].mean().nlargest(3)

# Create Excel with multiple sheets
wb = Workbook()

# Sheet 1: Content Performance
ws1 = wb.active
ws1.title = "Content Performance"

ws1['A1'] = "CONTENT TYPE PERFORMANCE BY PLATFORM"
ws1['A1'].font = Font(bold=True, size=12, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws1.merge_cells('A1:E1')

row = 2
headers = ['Platform', 'Content Type', 'Avg Engagement %', 'Max Engagement %', 'Total Followers']
for col_num, header in enumerate(headers, 1):
    ws1.cell(row=row, column=col_num).value = header
    ws1.cell(row=row, column=col_num).font = Font(bold=True, color="FFFFFF")
    ws1.cell(row=row, column=col_num).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

for row_num, (idx, row_data) in enumerate(content_performance.iterrows(), 3):
    ws1.cell(row=row_num, column=1).value = idx[0]
    ws1.cell(row=row_num, column=2).value = idx[1]
    ws1.cell(row=row_num, column=3).value = row_data['engagement_rate_mean']
    ws1.cell(row=row_num, column=4).value = row_data['engagement_rate_max']
    ws1.cell(row=row_num, column=5).value = int(row_data['followers_gained_sum'])

# Sheet 2: All posts with recommendations
ws2 = wb.create_sheet("Recommendations")

ws2['A1'] = "POST PERFORMANCE & AI RECOMMENDATIONS"
ws2['A1'].font = Font(bold=True, size=12, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
ws2.merge_cells('A1:H1')

row = 2
headers = ['Post ID', 'Platform', 'Content Type', 'Engagement %', 'Likes', 'Shares', 'Recommendation']
for col_num, header in enumerate(headers, 1):
    ws2.cell(row=row, column=col_num).value = header
    ws2.cell(row=row, column=col_num).font = Font(bold=True)

for row_num, (idx, row_data) in enumerate(df[['post_id', 'platform', 'content_type', 'engagement_rate', 'likes', 'shares', 'recommendation']].iterrows(), 3):
    for col_num, value in enumerate(row_data, 1):
        ws2.cell(row=row_num, column=col_num).value = value

# Sheet 3: Best posting times
ws3 = wb.create_sheet("Best Times")

ws3['A1'] = "BEST HOURS TO POST (by avg engagement)"
ws3['A1'].font = Font(bold=True, size=12, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
ws3.merge_cells('A1:B1')

ws3['A2'] = "Post Hour"
ws3['B2'] = "Avg Engagement %"
ws3['A2'].font = Font(bold=True)
ws3['B2'].font = Font(bold=True)

for row_num, (hour, engagement) in enumerate(best_hours.items(), 3):
    ws3.cell(row=row_num, column=1).value = f"{hour}:00"
    ws3.cell(row=row_num, column=2).value = round(engagement, 2)
    ws3.cell(row=row_num, column=2).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Adjust widths
for ws in [ws1, ws2, ws3]:
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

wb.save('content_optimizer.xlsx')
print("✅ Content optimization report generated: content_optimizer.xlsx")
