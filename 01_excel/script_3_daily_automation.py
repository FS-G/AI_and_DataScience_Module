"""
SCRIPT 3: Automated Daily Report Generator
Demonstrates: Date filtering, dynamic Excel generation, automatic file naming
This would run on a schedule (cron job or Windows Task Scheduler)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime

# Read data
df = pd.read_csv('social_media_data.csv')

# Filter for today's data (example: 2026-08-10)
today = '2026-08-10'
today_data = df[df['date'] == today]

if today_data.empty:
    print(f"No data for {today}")
else:
    # Calculate daily metrics
    total_engagement = today_data['engagement_rate'].sum()
    total_followers = today_data['followers_gained'].sum()
    top_post = today_data.loc[today_data['engagement_rate'].idxmax()]

    # Platform breakdown for today
    platform_daily = today_data.groupby('platform').agg({
        'likes': 'sum',
        'engagement_rate': 'mean',
        'followers_gained': 'sum'
    }).round(2)

    # Create automated report
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    # Header with date
    ws['A1'] = f"DAILY SOCIAL MEDIA REPORT - {today}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws.merge_cells('A1:D1')

    # KPIs
    ws['A3'] = "KEY METRICS"
    ws['A3'].font = Font(bold=True, size=11)

    kpi_data = [
        ('Total Posts', len(today_data)),
        ('Total Engagement %', round(total_engagement, 2)),
        ('Followers Gained', total_followers),
        ('Avg Engagement %', round(today_data['engagement_rate'].mean(), 2))
    ]

    for idx, (metric, value) in enumerate(kpi_data, 4):
        ws.cell(row=idx, column=1).value = metric
        ws.cell(row=idx, column=1).font = Font(bold=True)
        ws.cell(row=idx, column=2).value = value
        ws.cell(row=idx, column=2).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Top post highlight
    ws['A9'] = "🏆 TOP PERFORMING POST"
    ws['A9'].font = Font(bold=True, size=11, color="FFFFFF")
    ws['A9'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    ws['A10'] = f"Post ID: {top_post['post_id']}"
    ws['A11'] = f"Platform: {top_post['platform']}"
    ws['A12'] = f"Content Type: {top_post['content_type']}"
    ws['A13'] = f"Engagement Rate: {top_post['engagement_rate']}%"
    ws['A14'] = f"Likes: {top_post['likes']}"

    # Platform breakdown
    ws['A16'] = "PLATFORM BREAKDOWN"
    ws['A16'].font = Font(bold=True, size=11)

    ws['A17'] = "Platform"
    ws['B17'] = "Likes"
    ws['C17'] = "Avg Engagement"
    ws['D17'] = "Followers"

    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}17'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}17'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for row_num, (platform, row_data) in enumerate(platform_daily.iterrows(), 18):
        ws.cell(row=row_num, column=1).value = platform
        ws.cell(row=row_num, column=2).value = int(row_data['likes'])
        ws.cell(row=row_num, column=3).value = row_data['engagement_rate']
        ws.cell(row=row_num, column=4).value = int(row_data['followers_gained'])

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15

    # Save with timestamp
    filename = f"daily_report_{today}.xlsx"
    wb.save(filename)
    print(f"✅ Daily report saved: {filename}")
    print(f"📊 Summary: {total_followers} new followers, {round(total_engagement, 2)}% total engagement")
